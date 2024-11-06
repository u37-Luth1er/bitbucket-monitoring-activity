from openpyxl import Workbook
from datetime import datetime
from openpyxl import Workbook, load_workbook 
from created_repo import repo_created_monitoring
from updated_repo import repo_updated_monitoring
from webhook_google import Webhook
import pytz
from email.message import EmailMessage
import json, random, sys, requests, smtplib
import argparse

parser = argparse.ArgumentParser(description='Bitbucket Monitoring AppSec Tool')
parser.add_argument('-k','--apikey', type=str, required=True, help='--apikey BITBUCKET_API_KEY')
parser.add_argument('-u','--updated', type=str, required=False, help='--update DD/MM/AAAA')
parser.add_argument('-c','--created', type=str, required=False, help='--created DD/MM/AAAA')
parser.add_argument('-n','--notification', required=False, type=str,help='--notification')
parser.add_argument('-e', '--export', required=False, type=str ,help='--export')
args = parser.parse_args()


class Export_all_repositories:
    def __init__(self, token):
        self._token = token
    def monitoring(self):
        secret = { 'Authorization': 'Basic {}'.format(self._token),
                'Accept' : 'application/json'}
        workspace = '' # your workspace here
        base_url = "https://api.bitbucket.org/2.0"
        repos_list = []
        url = f"{base_url}/repositories/{workspace}"
        while url:
            response = requests.get(url, headers = secret)
            response_data = json.loads(response.content)
            repos_list += response_data["values"]
            if "next" in response_data:
                url = response_data["next"]
                url = None
            else:
                url = None

        workbook = Workbook()
        sheet = workbook.active
        try:
            workbook = load_workbook(filename="repositorios-total.xlsx")
            sheet = workbook.active
            existing_repos = {sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)}
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            existing_repos = set()
            header = ["name", "description", "url","project","created", "last udpate"]
            sheet.append(header)
        for repo in repos_list:
            created_on = datetime.strptime(repo["created_on"], '%Y-%m-%dT%H:%M:%S.%f%z').strftime('%d/%m/%Y')
            updated_on = datetime.strptime(repo["updated_on"], '%Y-%m-%dT%H:%M:%S.%f%z').strftime('%d/%m/%Y')
            if args.export:
                row = [
                    repo["name"],
                    repo.get("description", ""),
                    repo["links"]["html"]["href"],
                    repo["project"]["key"],
                    created_on,
                    updated_on
                ]
                sheet.append(row)
            workbook.save(filename="repositorios-total.xlsx")
            if args.notification:
                if repo["name"] not in existing_repos:
                    row = [
                        repo["name"],
                        repo.get("description", ""),
                        repo["links"]["html"]["href"],
                        repo["project"]["key"],
                        created_on,
                        updated_on
                    ]
                    content = '''

                '''.format(repo["name"], repo["links"]["html"]["href"], repo["project"]["key"], created_on, updated_on)
                    Webhook_instance = Webhook(content)
                    Webhook_instance.trigger_google_chats()
                    input('dbg')
                    sheet.append(row)
            workbook.save(filename="repositorios-total.xlsx") 

if args.created:
    Bitbucket_instance = repo_created_monitoring(args.apikey, args.created)
    Bitbucket_instance.created_monitoring()
elif args.updated:
    Bitbucket_instance = repo_updated_monitoring(args.apikey, args.updated)
    Bitbucket_instance.updated_monitoring()
elif args.export or args.notification:
    Bitbucket_instance = Export_all_repositories(args.apikey)
    Bitbucket_instance.monitoring()
else:
    print('nothing else error return')